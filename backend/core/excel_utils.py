"""
core/excel_utils.py

Professional Excel styling utilities.

Palette  : calm slate-blue / off-white  —  clean, readable, business-grade.
Font     : Arial throughout (professional, universally available).
Borders  : subtle inner grid, slightly heavier outer perimeter.
Alignment: auto-detected per value type (numbers right, dates centre, text left).
"""

from __future__ import annotations

from datetime import date, datetime

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Design tokens ─────────────────────────────────────────────────────────────

_P = {
    "header_bg":   "3A5F8A",   # slate blue   — calm, modern header band
    "header_fg":   "FFFFFF",   # white         — crisp on dark header
    "title_fg":    "2C4770",   # deep navy     — anchors the title hierarchy
    "subtitle_fg": "6B8DAE",   # muted slate   — softer than the title
    "alt_row":     "EFF4F9",   # pale ice-blue — barely-there alternating tint
    "border":      "C8D8E8",   # light steel   — present but not intrusive
    "body_fg":     "2D2D2D",   # near-black    — comfortable reading contrast
}

# Explicit number / date display formats
_FMT = {
    "date":       "YYYY-MM-DD",
    "datetime":   "YYYY-MM-DD HH:MM",
    "currency":   '"$"#,##0.00',
    "percentage": "0.0%",
    "integer":    "#,##0",
    "float":      "#,##0.00",
}

_FONT_BODY     = dict(name="Arial", size=10, color=_P["body_fg"])
_FONT_HEADER   = dict(name="Arial", size=10, bold=True,   color=_P["header_fg"])
_FONT_TITLE    = dict(name="Arial", size=13, bold=True,   color=_P["title_fg"])
_FONT_SUBTITLE = dict(name="Arial", size=10, italic=True, color=_P["subtitle_fg"])


# ── Internal helpers ──────────────────────────────────────────────────────────

def _side(weight: str = "thin", color: str | None = None) -> Side:
    return Side(style=weight, color=color or _P["border"])


def _cell_border(local_col: int, total_cols: int, is_last_row: bool) -> Border:
    """Thin inner grid with a medium perimeter on the table edges."""
    return Border(
        left   = _side("medium" if local_col == 1          else "thin"),
        right  = _side("medium" if local_col == total_cols else "thin"),
        top    = _side("thin"),
        bottom = _side("medium" if is_last_row             else "thin"),
    )


def _detect_format(value, hint: str | None = None) -> tuple[str | None, str]:
    """
    Return (number_format | None, horizontal_alignment).

    `hint` is an explicit key from _FMT: 'currency', 'percentage', etc.
    When hint is absent the type of value is used for inference.
    """
    if hint in _FMT:
        right_aligned = hint in ("currency", "percentage", "integer", "float")
        return _FMT[hint], ("right" if right_aligned else "center")
    if isinstance(value, bool):          # bool before int — bool IS int in Python
        return None, "center"
    if isinstance(value, datetime):
        return _FMT["datetime"], "center"
    if isinstance(value, date):
        return _FMT["date"], "center"
    if isinstance(value, int):
        return _FMT["integer"], "right"
    if isinstance(value, float):
        return _FMT["float"], "right"
    return None, "left"


def _col_display_width(values: list[str], cap: int = 52) -> float:
    """Compute a comfortable column width from displayed string lengths."""
    if not values:
        return 12
    longest = max(len(v) for v in values)
    return float(max(10, min(longest + 4, cap)))


# ── Public API ────────────────────────────────────────────────────────────────

def add_sheet_title(
    worksheet,
    title: str,
    subtitle: str | None = None,
    start_row: int = 1,
    col_span: int | None = None,
) -> int:
    """
    Write a styled title (and optional subtitle) at the top of a worksheet.
    Cells are merged across `col_span` columns for a polished look.

    Call this *before* writing headers / data so row numbers line up.
    Returns the next available row index after the title block.

    Usage
    -----
    next_row = add_sheet_title(ws, "Sales Report Q2 2026",
                               subtitle="Generated 2026-06-15",
                               col_span=ws.max_column)
    # … write header at next_row, data at next_row+1 …
    style_excel_worksheet(ws, header_row=next_row)
    """
    max_col  = col_span or worksheet.max_column or 1
    next_row = start_row

    def _write(row, text, font_kw, height):
        cell = worksheet.cell(row=row, column=1)
        cell.value     = text
        cell.font      = Font(**font_kw)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if max_col > 1:
            worksheet.merge_cells(
                start_row=row, start_column=1,
                end_row=row,   end_column=max_col,
            )
        worksheet.row_dimensions[row].height = height

    _write(next_row, title, _FONT_TITLE, 30)
    next_row += 1

    if subtitle:
        _write(next_row, subtitle, _FONT_SUBTITLE, 18)
        next_row += 1

    # Thin visual spacer row between title block and header band
    worksheet.row_dimensions[next_row].height = 5
    next_row += 1

    return next_row


def apply_column_color_scale(
    worksheet,
    col_idx: int,
    data_start_row: int,
    data_end_row: int,
    low_color:  str = "FDECEA",   # soft rose  — low values
    mid_color:  str = "FFFFFF",   # white      — midpoint
    high_color: str = "D6EDD6",   # sage green — high values
) -> None:
    """
    Apply a gentle 3-colour scale to a numeric column.
    Only use this where the gradient genuinely aids interpretation
    (e.g. a performance score or revenue column), not decoratively.

    Usage
    -----
    apply_column_color_scale(ws, col_idx=4,
                             data_start_row=2, data_end_row=50)
    """
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
    worksheet.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="min",        start_color=low_color,
            mid_type="percentile",   mid_value=50,  mid_color=mid_color,
            end_type="max",          end_color=high_color,
        ),
    )


def style_excel_worksheet(
    worksheet,
    header_row: int = 1,
    has_header: bool = True,
    alternate_rows: bool = True,
    freeze_panes: bool = True,
    freeze_first_column: bool = False,
    add_filters: bool = True,
    column_formats: dict | None = None,
) -> None:
    """
    Apply professional, calm styling to an openpyxl Worksheet in-place.

    Parameters
    ----------
    worksheet            : openpyxl Worksheet to style.
    header_row           : 1-based row index of the column-header row.
    has_header           : style the header row with a polished dark band.
    alternate_rows       : alternate pale-blue / white row tints in data rows.
    freeze_panes         : freeze the header row so it stays visible on scroll.
    freeze_first_column  : also freeze column A (useful for ID / name columns).
    add_filters          : add Excel AutoFilter to the header row.
    column_formats       : optional dict of {1-based column index → format key}.
                           Valid keys: 'currency', 'percentage', 'date',
                                       'datetime', 'integer', 'float'.
                           Values not in this dict are auto-detected from type.
                           Example: {3: 'currency', 5: 'percentage', 7: 'date'}

    Styling decisions
    -----------------
    - Header   : slate-blue fill (#3A5F8A), white bold Arial 10, centred,
                 medium outer border, thin inner border.
    - Data rows: Arial 10 near-black; numbers right-aligned, dates centred,
                 text left-aligned; medium outer perimeter, thin inner grid.
    - Alternate : every other data row gets a pale ice-blue tint (#EFF4F9).
    - Long text  : strings > 40 chars get wrap_text=True; row height expands.
    - Widths     : auto-sized from content (10 – 52 chars, +4 padding).
    """
    column_formats = column_formats or {}

    min_col    = worksheet.min_column or 1
    max_col    = worksheet.max_column or 1
    max_row    = worksheet.max_row    or 1
    total_cols = max_col - min_col + 1

    # ── Header row ───────────────────────────────────────────────────────────
    if has_header:
        worksheet.row_dimensions[header_row].height = 32
        for cell in worksheet[header_row]:
            col_local = cell.column - min_col + 1
            cell.font      = Font(**_FONT_HEADER)
            cell.fill      = PatternFill("solid", fgColor=_P["header_bg"])
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left   = _side("medium" if col_local == 1          else "thin"),
                right  = _side("medium" if col_local == total_cols else "thin"),
                top    = _side("medium"),
                bottom = _side("medium"),
            )

    # ── Data rows ────────────────────────────────────────────────────────────
    data_start = (header_row + 1) if has_header else header_row

    for row_cells in worksheet.iter_rows(min_row=data_start, max_row=max_row):
        row_idx     = row_cells[0].row
        is_alt      = alternate_rows and (row_idx - data_start) % 2 == 1
        is_last_row = row_idx == max_row
        row_fill    = PatternFill("solid", fgColor=_P["alt_row"]) if is_alt else None
        row_height  = 18  # default; may grow for wrapped text

        for cell in row_cells:
            col_local = cell.column - min_col + 1
            hint      = column_formats.get(cell.column)
            num_fmt, halign = _detect_format(cell.value, hint)

            if num_fmt:
                cell.number_format = num_fmt

            needs_wrap = (
                halign == "left"
                and isinstance(cell.value, str)
                and len(cell.value) > 40
            )
            if needs_wrap:
                row_height = max(row_height, 36)

            cell.font      = Font(**_FONT_BODY)
            cell.alignment = Alignment(
                horizontal=halign,
                vertical="center",
                wrap_text=needs_wrap,
            )
            if row_fill:
                cell.fill = row_fill

            cell.border = _cell_border(col_local, total_cols, is_last_row)

        worksheet.row_dimensions[row_idx].height = row_height

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        str_values = [
            str(worksheet.cell(row=r, column=col_idx).value)
            for r in range(1, max_row + 1)
            if worksheet.cell(row=r, column=col_idx).value is not None
        ]
        worksheet.column_dimensions[col_letter].width = _col_display_width(str_values)

    # ── Freeze panes ─────────────────────────────────────────────────────────
    if freeze_panes and has_header:
        freeze_row = header_row + 1
        freeze_col = (min_col + 1) if freeze_first_column else min_col
        worksheet.freeze_panes = worksheet.cell(
            row=freeze_row, column=freeze_col
        ).coordinate

    # ── AutoFilter ───────────────────────────────────────────────────────────
    if add_filters and has_header:
        first = worksheet.cell(row=header_row, column=min_col).coordinate
        last  = worksheet.cell(row=header_row, column=max_col).coordinate
        worksheet.auto_filter.ref = f"{first}:{last}"